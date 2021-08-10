# -*- coding: utf-8 -*-
"""
Created on Sun Oct 25 18:50:03 2020

@author: Hp
"""

#-------------------------------------------------------------------------------
# Name:        module2
# Purpose:
#
# Author:      Sureshkumar
#
# Created:     24/07/2020
# Copyright:   (c) dsk 2020
# Licence:     <your licence>
#-------------------------------------------------------------------------------

import imaplib
import sys
import email
import os
import datetime
from datetime import timedelta
from copy import copy
import openpyxl
from openpyxl.utils import get_column_letter
import pandas as pd
import csv
import shutil
import calendar
from time import strptime
from pandas.tseries.offsets import MonthEnd,MonthBegin

def save_workbook_excel_file(workbook , filename: str):
    """Tries to save created data to excel file"""
    try:
        workbook.save(filename)
    except PermissionError:
        print("Error: No permission to save file.")
        
def invert_row_column_week(filename: str):
    print("Transpose Weekly Data")
    workbook = openpyxl.load_workbook(filename,data_only=True)
    sheet_names = workbook.sheetnames
    sheet = workbook[sheet_names[1]]
    #sheet.delete_rows(14, 1)
    #sheet.delete_rows(26, 5)
    #sheet.delete_rows(29, 1)
    #workbook.create_sheet(index=0, title='tmp_sheet')
    tmp_sheet = workbook['tmp_sheet']
    data = []
    for row in sheet:
        cells = []
        for cell in row:
            cells.append(cell)
        data.append(cells)

    for x in range(0, len(data)):
        for y in range(0, len(data[x])):
            column = get_column_letter(x + 1)
            row = str(y + 1)
            tmp_sheet[column + row] = copy(data[x][y].value)

    sheet_name = sheet.title
    del workbook[sheet_name]
    del workbook['Actual-Forecast 2020']
    tmp_sheet.title = sheet_name
    save_workbook_excel_file(workbook,'week_B2C_Revenues_2020.xlsx' )
    print("Done Transpose Weekly Data and writing in seperate buffer file")    
    return 1
    
def main():
    print("start")
    print("fetch revenue xlsx file and split them into Weekly and Monthly")
    file_name ='Intigral B2C Revenues  2020 W43.xlsx'
    invert_row_column_week(file_name)
    print("Done Transpose Weekly&Monthly Data and writing in seperate buffer file")
    print("end")